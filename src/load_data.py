from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal, TypeAlias

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


Dia: TypeAlias = Literal[0, 7, 14]
Concentracao: TypeAlias = Literal[0, 25, 50, 75, 100, -1]
Sistema: TypeAlias = Literal["AI", "A", "BI", "B", "CN", "CP"]

CONCENTRACOES_PADRAO: list[Concentracao] = [0, 25, 50, 75, 100]
DIAS_PADRAO: list[Dia] = [0, 7, 14]


@dataclass
class Amostra:
    fungo: str
    sistema: Sistema
    dia: Dia
    concentracao: Concentracao
    analise: str
    valor: Any


class Extrator:
    def __init__(self, file_path: str | Path, sheet: str, fungo: str = "") -> None:
        self.file_path = Path(file_path)
        self.sheet_name = sheet
        self.fungo = fungo or self.file_path.stem
        self._amostras: list[Amostra] = []
        self.error_message = ""

        print(f"Iniciando extrator: {self.fungo} | planilha: {self.sheet_name}")

        if not self.file_path.exists():
            self.error_message = f"ERRO - Caminho {self.file_path} não existe"
            print(self.error_message)
            return

        self.workbook = load_workbook(self.file_path, data_only=True)
        self.worksheet = self.workbook[self.sheet_name]

    @property
    def amostras(self) -> list[Amostra]:
        return self._amostras

    def extrair_ufc(self) -> list[Amostra]:
        return []

    def extrair_base_seca(self) -> list[Amostra]:
        if self.error_message:
            return []

        configuracoes = [
            ("AI", 3),
            ("A", 13),
            ("BI", 23),
            ("B", 33),
        ]

        for sistema, start_row in configuracoes:
            self._extrair_linhas_fixas(
                start_col_letter="g",
                start_row=start_row,
                row_step=2,
                total_rows=5,
                dia=14,
                sistema=sistema,
                analise="Massa Fungica",
                concentracoes=CONCENTRACOES_PADRAO,
            )

        return self._amostras

    def extrair_analise_macroscopica(self, dia: Dia) -> list[Amostra]:
        if self.error_message:
            return []

        analises_por_coluna = {
            "d": "Sementes Germinadas",
            "e": "TG",
            "f": "GRS",
            "g": "IG",
            "h": "CRR",
            "i": "IGN",
            "j": "IER",
        }

        sistemas_com_bloco = [
            ("AI", 4),
            ("A", 9),
            ("BI", 14),
            ("B", 19),
        ]

        controles = [
            ("CN", 24),
            ("CP", 25),
        ]

        for col_letter, analise in analises_por_coluna.items():
            for sistema, start_row in sistemas_com_bloco:
                self.extrair_coluna(
                    start_col_letter=col_letter,
                    start_row=start_row,
                    dia=dia,
                    sistema=sistema,
                    analise=analise,
                )

            for sistema, row in controles:
                self.extrair_celula(
                    start_col_letter=col_letter,
                    start_row=row,
                    dia=dia,
                    sistema=sistema,
                    analise=analise,
                    concentracao=-1,
                )

        for amostra in self._amostras:
            if amostra.valor == "#DIV/0!":
                amostra.valor = 0

        return self._amostras

    def extrair_metais_toxicos(self) -> list[Amostra]:
        if self.error_message:
            return []

        configuracoes = [
            (0, "i"),
            (14, "o"),
        ]

        sistemas = [
            ("AI", 5),
            ("A", 10),
            ("BI", 15),
            ("B", 20),
        ]

        for dia, col_letter in configuracoes:
            for sistema, start_row in sistemas:
                self.extrair_coluna(
                    start_col_letter=col_letter,
                    start_row=start_row,
                    dia=dia,
                    sistema=sistema,
                    analise="Zn",
                )

        return self._amostras

    def extrair_fisicoquimico(self) -> list[Amostra]:
        if self.error_message:
            return []

        configuracoes = [
            ("d", 5, "AI", "pH"),
            ("d", 11, "A", "pH"),
            ("d", 17, "BI", "pH"),
            ("d", 23, "B", "pH"),
            ("h", 5, "AI", "Turbidez"),
            ("h", 11, "A", "Turbidez"),
            ("h", 17, "BI", "Turbidez"),
            ("h", 23, "B", "Turbidez"),
            ("l", 5, "AI", "Condutividade"),
            ("l", 11, "A", "Condutividade"),
            ("l", 17, "BI", "Condutividade"),
            ("l", 23, "B", "Condutividade"),
            ("p", 5, "AI", "Cor"),
            ("p", 11, "A", "Cor"),
            ("p", 17, "BI", "Cor"),
            ("p", 23, "B", "Cor"),
            ("t", 5, "AI", "DQO"),
            ("t", 11, "A", "DQO"),
            ("t", 17, "BI", "DQO"),
            ("t", 23, "B", "DQO"),
        ]

        for col_letter, row, sistema, analise in configuracoes:
            self.extrair_quadrante(
                start_col_letter=col_letter,
                start_row=row,
                sistema=sistema,
                analise=analise,
            )

        return self._amostras

    def extrair_celula(
        self,
        start_col_letter: str,
        start_row: int,
        dia: Dia,
        sistema: Sistema,
        analise: str,
        concentracao: Concentracao,
    ) -> None:
        print(
            f"Extraindo célula | sistema={sistema} | análise={analise} "
            f"| fungo={self.fungo} | concentração={concentracao}"
        )

        for row in self._iterar_intervalo(
            start_col_letter=start_col_letter,
            start_row=start_row,
            ncols=1,
            nrows=1,
        ):
            for cell in row:
                self._adicionar_amostra(
                    sistema=sistema,
                    dia=dia,
                    concentracao=concentracao,
                    analise=analise,
                    valor=cell.value,
                )

    def extrair_coluna(
        self,
        start_col_letter: str,
        start_row: int,
        dia: Dia,
        sistema: Sistema,
        analise: str,
    ) -> None:
        print(
            f"Extraindo coluna | sistema={sistema} | análise={analise} | fungo={self.fungo}"
        )

        for row_index, row in enumerate(
            self._iterar_intervalo(
                start_col_letter=start_col_letter,
                start_row=start_row,
                ncols=1,
                nrows=5,
            )
        ):
            concentracao = CONCENTRACOES_PADRAO[row_index]
            for cell in row:
                self._adicionar_amostra(
                    sistema=sistema,
                    dia=dia,
                    concentracao=concentracao,
                    analise=analise,
                    valor=cell.value,
                )

    def extrair_quadrante(
        self,
        start_col_letter: str,
        start_row: int,
        sistema: Sistema,
        analise: str,
    ) -> None:
        print(
            f"Extraindo quadrante | sistema={sistema} | análise={analise} | fungo={self.fungo}"
        )

        for row_index, row in enumerate(
            self._iterar_intervalo(
                start_col_letter=start_col_letter,
                start_row=start_row,
                ncols=3,
                nrows=5,
            )
        ):
            concentracao = CONCENTRACOES_PADRAO[row_index]

            for col_index, cell in enumerate(row):
                dia = DIAS_PADRAO[col_index]
                self._adicionar_amostra(
                    sistema=sistema,
                    dia=dia,
                    concentracao=concentracao,
                    analise=analise,
                    valor=cell.value,
                )

    def _extrair_linhas_fixas(
        self,
        start_col_letter: str,
        start_row: int,
        row_step: int,
        total_rows: int,
        dia: Dia,
        sistema: Sistema,
        analise: str,
        concentracoes: list[Concentracao],
    ) -> None:
        for index in range(total_rows):
            row = start_row + index * row_step
            concentracao = concentracoes[index]
            self.extrair_celula(
                start_col_letter=start_col_letter,
                start_row=row,
                dia=dia,
                sistema=sistema,
                analise=analise,
                concentracao=concentracao,
            )

    def _iterar_intervalo(
        self,
        start_col_letter: str,
        start_row: int,
        ncols: int,
        nrows: int,
    ):
        start_col = column_index_from_string(start_col_letter)
        end_col = start_col + ncols - 1
        end_row = start_row + nrows - 1
        end_col_letter = get_column_letter(end_col)

        cell_range = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        return self.worksheet[cell_range]

    def _adicionar_amostra(
        self,
        sistema: Sistema,
        dia: Dia,
        concentracao: Concentracao,
        analise: str,
        valor: Any,
    ) -> None:
        if valor is None or valor == "":
            valor = np.nan

        amostra = Amostra(
            fungo=self.fungo,
            sistema=sistema,
            dia=dia,
            concentracao=concentracao,
            analise=analise,
            valor=valor,
        )
        self._amostras.append(amostra)